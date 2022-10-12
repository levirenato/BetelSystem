from openpyxl import Workbook, load_workbook
import datetime
import os
import Impress
import sendEmail

# database with the products
bd_inj_vol = {
    '15,7g': 24360,
    '83g': 3200,
    '700g': 500,
    '650g': 500,
    'Tampa 5L': 2000,
    'Tampa 20L': 2000,
    'Alça': 2000,
}

# sheets in use
workbook = load_workbook('BaseInjecao.xlsx')
wb = load_workbook('BaseSopro.xlsx')
injecao = workbook["Injeção"]
sopro = wb["Sopro"]
# variables
M_numero = '' #is a global var to put the name in the file
data = datetime.datetime.today().strftime('%d-%m-%y') #is a global var to put the name in the file
M_cor = ''  #is a global var to put the name in the file

# function to get a value and fill in sheet

#Function to 'InjecaoBase' sheet
def injecao_title(value):
    global M_numero
    M_numero = value
    inj_title = f'Programação Da Máquina {M_numero}'
    injecao['F4'] = inj_title
    os.system('cls')


def injecao_prod(valor):
    inj_prod = valor
    injecao['F9'] = inj_prod
    injecao['M9'] = inj_prod


def injecao_cor(valor):
    global M_cor
    M_cor = valor.title()
    inj_cor = M_cor
    injecao['G9'] = inj_cor
    injecao['N9'] = inj_cor


def injecao_data():
    data = datetime.datetime.today().strftime('%d/%m/%y')
    injecao['J6'] = f'Data: {data}'

#that function get the amount will be produced, and return the total of boxes and the bulk in a box
def injecao_qnt_total(quantidade,produto,vol):
    input_inj_qnt = quantidade
    # Find the poduct's size box and retur the total boxes will be used
    inj_qnt = input_inj_qnt
    verifica = bd_inj_vol.get(produto)
    injecao['H9'] = f'{inj_qnt} ({int(inj_qnt / verifica)} {vol})'
    injecao['O9'] = f'{bd_inj_vol.get(produto)} ({vol})'


def injecao_cli(valor):
    input_inj_cli = valor.title()
    inj_cli = input_inj_cli
    injecao['K9'] = inj_cli


def injecao_COD(valor):
    input_inj_COD = valor
    inj_COD = input_inj_COD
    injecao['P9'] = inj_COD


def injecao_OBs(valor):
    input_inj_Obs = valor
    inj_Obs = input_inj_Obs
    injecao['Q9'] = inj_Obs

#in here the name of file is created following Number of Machine - Date - Color
def start():
    Nome = f'M{M_numero} {data} {M_cor}'
    workbook.save(f"Historico/{Nome}.xlsx")
    print(f'Arquivo {Nome} salvo!')

#the function find and get the file created to send email function in the script
def send_email(email):
    Nome = f'M{M_numero} {data} {M_cor}'
    anexo = f"Historico/{Nome}.xlsx"
    sendEmail.send_email(anexo,email)
#its the same but to impress
def imprimir():
    Nome = f'M{M_numero} {data} {M_cor}'
    anexo_nome = f'{Nome}.xlsx'
    Impress.imprimir(anexo_nome)

#################################      SOPRO        #######################################################

#The same structure but to another departament
sM_numero = ''
sM_cor = ''
def sopro_prod(valor):
    s_prod = valor
    sopro['B8'] = s_prod
    sopro['I8'] = s_prod


def sopro_title(v):
    global sM_numero
    sM_numero = v
    s_title = f'Programação Da Máquina {sM_numero} ({sopro["B8"].value}) Sopro'
    sopro['B3'] = s_title


def sopro_cor(valor):
    global sM_cor
    sM_cor = valor.title()
    sopro['C8'] = sM_cor
    sopro['J8'] = sM_cor


def sopro_data():
    data = datetime.datetime.today().strftime('%d/%m/%y')
    sopro['F5'] = f'Data: {data}'


def sopro_qnt_total(quantidade):
    sopro['D8'] = quantidade


def sopro_cli(valor):
    sopro_cli = valor.title()
    sopro['F8'] = sopro_cli



def sopro_OBs(valor):
    sopro_Obs = valor
    sopro['L8'] = sopro_Obs


def sstart():
    sNome = f'sopro M{sM_numero} {data} {sM_cor}'
    wb.save(f"Historico/{sNome}.xlsx")
    print(f'Arquivo {sNome} salvo!')


def ssend_email(email):
    sNome = f'sopro M{sM_numero} {data} {sM_cor}'
    anexo = f"Historico/{sNome}.xlsx"
    sendEmail.send_email(anexo,email)
def simprimir():
    sNome = f'sopro M{sM_numero} {data} {sM_cor}'
    anexo_nome = f'{sNome}.xlsx'
    Impress.imprimir(anexo_nome)

